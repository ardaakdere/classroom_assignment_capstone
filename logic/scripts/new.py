from openpyxl import load_workbook
from logic import models
from datetime import datetime

def run():
    # Load the Excel workbook
    wb = load_workbook('/Users/ardaakdere/Desktop/ClassroomAssignmentSystem/classroom_assignment/logic/scripts/Capstone.XLSX')
    ws = wb['beşiktaş_tüm_dersler']

    # Create a 2D array to store the data
    data = []

    # # # Iterate through all rows in the worksheet
    for row in ws.iter_rows(values_only=True):
        # Append the row data as a list to the 2D array
        data.append(list(row))

    # Print the data
    for row in data[1:]:
        print('DONE!')
        classroom = models.Classroom(classroom_number = row[2], classroom_quota = 100)
        classroom.save()
        start_time = datetime(2000, 10, 10, row[4].hour, row[4].minute)
        end_time = datetime(2000, 10, 10, row[5].hour, row[5].minute)
        course = models.Course(course_name = row[1], course_code = row[0], classroom = classroom, students_quota = row[6], start_date = start_time, end_date = end_time)
        course.save()
